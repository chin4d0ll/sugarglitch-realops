#!/usr/bin/env python3
"""
Step 2: Data Export Tools
Export DM data in multiple formats (JSON, CSV, PDF, Excel)
"""

import sqlite3
import json
import csv
import pandas as pd
from datetime import datetime
from pathlib import Path

try:
    from fpdf import FPDF
except ImportError:
    import subprocess
    import sys
    print("Installing fpdf2...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "fpdf2"])
    from fpdf import FPDF

try:
    import openpyxl
except ImportError:
    import subprocess
    import sys
    print("Installing openpyxl...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

class DataExporter:
    def __init__(self, db_path="/workspaces/sugarglitch-realops/advanced_dm_database_1748742706.sqlite"):
        self.db_path = db_path
        self.export_dir = Path("/workspaces/sugarglitch-realops/exports")
        self.export_dir.mkdir(exist_ok=True)
        
    def get_database_data(self):
        """Retrieve all data from database"""
        conn = sqlite3.connect(self.db_path)
        
        # Get extraction sessions
        sessions_df = pd.read_sql_query("SELECT * FROM extraction_sessions", conn)
        
        # Get threads
        threads_df = pd.read_sql_query("SELECT * FROM dm_threads", conn)
        
        # Get messages
        messages_df = pd.read_sql_query("SELECT * FROM dm_messages", conn)
        
        conn.close()
        
        return {
            'sessions': sessions_df,
            'threads': threads_df,
            'messages': messages_df
        }
    
    def export_to_json(self):
        """Export all data to JSON format"""
        print("📄 Exporting to JSON...")
        
        data = self.get_database_data()
        
        # Convert DataFrames to dictionaries
        export_data = {
            'export_info': {
                'timestamp': datetime.now().isoformat(),
                'total_sessions': len(data['sessions']),
                'total_threads': len(data['threads']),
                'total_messages': len(data['messages'])
            },
            'extraction_sessions': data['sessions'].to_dict('records'),
            'dm_threads': data['threads'].to_dict('records'),
            'dm_messages': data['messages'].to_dict('records')
        }
        
        # Save to JSON file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        json_file = self.export_dir / f"dm_data_export_{timestamp}.json"
        
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False)
        
        print(f"✅ JSON export saved: {json_file}")
        return json_file
    
    def export_to_csv(self):
        """Export data to CSV files (one per table)"""
        print("📊 Exporting to CSV...")
        
        data = self.get_database_data()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        csv_files = []
        
        # Export each table to separate CSV
        for table_name, df in data.items():
            csv_file = self.export_dir / f"dm_{table_name}_{timestamp}.csv"
            df.to_csv(csv_file, index=False, encoding='utf-8')
            csv_files.append(csv_file)
            print(f"✅ CSV saved: {csv_file}")
        
        return csv_files
    
    def export_to_excel(self):
        """Export all data to Excel with multiple sheets"""
        print("📈 Exporting to Excel...")
        
        data = self.get_database_data()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = self.export_dir / f"dm_data_complete_{timestamp}.xlsx"
        
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # Write each DataFrame to a separate sheet
            data['sessions'].to_excel(writer, sheet_name='Extraction_Sessions', index=False)
            data['threads'].to_excel(writer, sheet_name='DM_Threads', index=False)
            data['messages'].to_excel(writer, sheet_name='DM_Messages', index=False)
            
            # Create summary sheet
            summary_data = {
                'Metric': [
                    'Total Extraction Sessions',
                    'Total DM Threads', 
                    'Total Messages',
                    'Export Date'
                ],
                'Value': [
                    len(data['sessions']),
                    len(data['threads']),
                    len(data['messages']),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        print(f"✅ Excel export saved: {excel_file}")
        return excel_file
    
    def export_to_pdf_report(self):
        """Create a comprehensive PDF report"""
        print("📑 Creating PDF report...")
        
        data = self.get_database_data()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        pdf_file = self.export_dir / f"dm_analysis_report_{timestamp}.pdf"
        
        # Create PDF
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font('Arial', 'B', 16)
        
        # Title
        pdf.cell(0, 10, 'Instagram DM Analysis Report', 0, 1, 'C')
        pdf.ln(10)
        
        # Summary statistics
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'Summary Statistics', 0, 1)
        pdf.set_font('Arial', '', 12)
        
        pdf.cell(0, 8, f'Total Extraction Sessions: {len(data["sessions"])}', 0, 1)
        pdf.cell(0, 8, f'Total DM Threads: {len(data["threads"])}', 0, 1)
        pdf.cell(0, 8, f'Total Messages: {len(data["messages"])}', 0, 1)
        pdf.cell(0, 8, f'Report Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', 0, 1)
        pdf.ln(10)
        
        # Thread analysis
        if not data['threads'].empty:
            pdf.set_font('Arial', 'B', 14)
            pdf.cell(0, 10, 'Thread Analysis', 0, 1)
            pdf.set_font('Arial', '', 12)
            
            for idx, thread in data['threads'].head(10).iterrows():
                pdf.cell(0, 6, f'Thread: {thread["thread_id"]} - {thread["participants"]}', 0, 1)
                pdf.cell(0, 6, f'  Messages: {thread["message_count"]} | Last Activity: {thread["last_activity"]}', 0, 1)
                pdf.ln(2)
        
        pdf.output(str(pdf_file))
        print(f"✅ PDF report saved: {pdf_file}")
        return pdf_file
    
    def create_conversation_exports(self):
        """Create individual conversation exports"""
        print("💬 Creating individual conversation exports...")
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Get all threads
        cursor.execute("SELECT thread_id, participants, thread_title FROM dm_threads")
        threads = cursor.fetchall()
        
        conversation_files = []
        
        for thread_id, participants, thread_title in threads:
            # Get messages for this thread
            cursor.execute("""
                SELECT username, content, timestamp, message_type 
                FROM dm_messages 
                WHERE thread_id = ? 
                ORDER BY timestamp
            """, (thread_id,))
            
            messages = cursor.fetchall()
            
            if messages:
                # Create conversation export
                conversation_data = {
                    'thread_info': {
                        'thread_id': thread_id,
                        'participants': participants,
                        'thread_title': thread_title,
                        'message_count': len(messages)
                    },
                    'messages': [
                        {
                            'sender': msg[0],
                            'content': msg[1],
                            'timestamp': msg[2],
                            'type': msg[3]
                        }
                        for msg in messages
                    ]
                }
                
                # Save as JSON
                safe_name = "".join(c for c in thread_id if c.isalnum() or c in ('-', '_'))
                conv_file = self.export_dir / f"conversation_{safe_name}.json"
                
                with open(conv_file, 'w', encoding='utf-8') as f:
                    json.dump(conversation_data, f, indent=2, ensure_ascii=False)
                
                conversation_files.append(conv_file)
                print(f"✅ Conversation saved: {conv_file}")
        
        conn.close()
        return conversation_files
    
    def run_all_exports(self):
        """Run all export formats"""
        print("🚀 Starting complete data export...")
        print("=" * 50)
        
        exports = {}
        
        try:
            exports['json'] = self.export_to_json()
            exports['csv'] = self.export_to_csv()
            exports['excel'] = self.export_to_excel()
            exports['pdf'] = self.export_to_pdf_report()
            exports['conversations'] = self.create_conversation_exports()
            
            print("\n" + "=" * 50)
            print("📦 EXPORT COMPLETE!")
            print("=" * 50)
            print(f"📁 All files saved to: {self.export_dir}")
            print(f"📄 JSON: {exports['json'].name}")
            print(f"📊 CSV files: {len(exports['csv'])} files")
            print(f"📈 Excel: {exports['excel'].name}")
            print(f"📑 PDF: {exports['pdf'].name}")
            print(f"💬 Conversations: {len(exports['conversations'])} files")
            
            return exports
            
        except Exception as e:
            print(f"❌ Export error: {e}")
            return None

if __name__ == "__main__":
    print("🎯 Step 2: Data Export Tools")
    print("=" * 40)
    
    exporter = DataExporter()
    exports = exporter.run_all_exports()
    
    if exports:
        print("\n✅ All exports completed successfully!")
    else:
        print("\n❌ Export failed!")
